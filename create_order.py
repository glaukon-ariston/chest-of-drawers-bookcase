import openpyxl
import csv
import argparse
import datetime
import os
from order.PII import CUSTOMER_NAME, CUSTOMER_PHONE, CUSTOMER_EMAIL, CUSTOMER_ADDRESS


"""
Creates an Iverpan order document from a CSV cut list.

This script reads a CSV file containing a cut list and populates an Excel template
with the data to create an order document for the Iverpan cutting service.

Example:
python create_order.py --model-id "H2300xW600xD230_Mm19_Ms12" --service iverpan --template "order/template/iverpan_tablica_za_narudzbu.xlsx"
python create_order.py --model-id "H2300xW600xD230_Mm19_Ms12" --service elgrad --template "order/template/elgrad_tablica_za_narudzbu.xlsx"
python create_order.py --model-id "H2300xW600xD230_Mm18_Ms12" --service furnir --template "order/template/furnir_tablica_za_narudzbu.xlsx"
python create_order.py --model-id "H2300xW600xD230_Mm18_Ms12" --service sizekupres --template "order/template/sizekupres_tablica_za_narudzbu.xlsx"
"""

# Set up argument parser
parser = argparse.ArgumentParser(description='Create an Iverpan order document from a CSV cut list.')
parser.add_argument('--model-id', required=True, help='Model identifier.')
parser.add_argument('--service', required=True, help='Service id: iverpan | elgrad')
parser.add_argument('--template', required=True, help='Path to the Excel template file.')
args = parser.parse_args()

ORDER_DIR = f'order/export/{args.model_id}'

MATERIAL = {
    'iverpan': {
        # Iveral material
        # 58	HRAST PRIRODNI (nova struktura)	37307AN						19				08/22      						2/28       			203837307AN
        # 58	HRAST PRIRODNI (nova struktura)	37307AN						19				081937307AN						202537307AN			203837307AN
        'iveral': {
            'MEL-19': '37307AN',
            'MEL-12': '1615FH',
            'HDF-3': 'HDF Hrast',
        },

        # Iveral banding
        # 											                    05/22		    1/22		    2/22	    08/28	1/28	2/28	    08/42	1/42	2/42
        # 1	BIJELI FH FUNDER	1615FH		10				19	25		B05191615FH		B10191615FH		B20191615FH	B08251615FH		B20251615FH	B08381615FH		B20381615FH
        'edge_banding': {
            'MEL-19': '202537307AN',
            'MEL-12': 'B10191615FH',
        },
    },
    'elgrad': {
        # Elgrad material
        # H1145 Hrast Bardolino natur
        # W960 Klasična bijela 
        'iveral': {
            'MEL-19': 'H1145',      # Hrast Bardolino natur
            'MEL-12': 'H1145',      # Klasična bijela 
            'HDF-3': 'HDF-3-Hrast',
        },

        # Elgrad banding
        'edge_banding': {
            'MEL-19': '2',
            'MEL-12': '2',
        },
    },
    'furnir': {
        'iveral': {
            'MEL-19': '11145501',      # 18 MM. 5501 SN Slavonski hrast
            'MEL-12': '1116012',       # BIJELI 12 MM. 284/183 - 116 A BS
            'HDF-3': 'HDF-3-Hrast',
        },
        'edge_banding': {
            'MEL-19': '11145501',
            'MEL-12': '1116012',
        },
    },
    'sizekupres': {
        'iveral': {
            'MEL-19': '0135419  34140 RV/RV 18mm HRAST SANREMO CLASSIC',      # 0135419  34140 RV/RV 18mm HRAST SANREMO CLASSIC
            'MEL-12': '0111103  1615 SF 12mm BIJELI',       # BIJELI 12 MM. 284/183 - 116 A BS
            'HDF-3': 'HDF-3-Hrast',
        },
        'edge_banding': {
            'MEL-19': 'ABS 34140 RV 23/2',      # ABS 34140 RV 23/0,8 ; ABS 34140 RV 23/2
            'MEL-12': 'ABS 1615 SF 23/0,45',    # ABS 1615 SF 23/0,45 ; ABS 1615 SF 23/2 
        },
    },
}

CSV_HEADER = "material code,material thickness,dimension A (along wood grain),dimension B,count,edge banding A-1,edge banding A-2,edge banding B-1,edge banding B-2,panel name,panel description,cnc face holes,cnc side holes".split(',')


def verify_header(header, expected_header):
    # Check if the header starts with the expected header
    if not all(h_found == h_expected for h_found, h_expected in zip(header, expected_header)):
        print("Warning: headers do not match; there might be some additional columns beyond the expected header.")
        print("Expected prefix:", expected_header)
        print("Found:", header)
        exit(1)
    return True


def process_elgrad_order_with_material(service, material_code, thickness, workbook_file, csv_file, output_file):
    print(f"Creating order file: {output_file}")

    workbook = openpyxl.load_workbook(workbook_file)
    # Select the 'Elementi' worksheet
    sheet = workbook['Elementi']

    # Verify header
    header_row = sheet[3]
    header_values = [cell.value for cell in header_row]
    expected_header = ['R.B.',	'DULJINA', 'ŠIRINA', 'BR.KOM.',	'DUŽA MJERA',  'KRAĆA MJERA', 'DUŽA MJERA', 'KRAĆA MJERA', None]
    verify_header(header_values, expected_header)

    sheet.cell(row=1, column=2).value = f"{MATERIAL[service]['iveral'][material_code]} {thickness}mm"
    sheet.cell(row=1, column=7).value = f"{MATERIAL[service]['iveral'][material_code]}"
    sheet.cell(row=2, column=2).value = "da" if material_code != 'HDF-3' else "ne"

    # Read data from CSV and populate the Excel sheet
    with open(csv_file, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        
        # Read header row
        header_row = next(reader)
        verify_header(header_row, CSV_HEADER)

        # Start writing from row 14
        row_num = 4
        for row in reader:
            material_code2 = row[0]
            thickness2 = row[1]
            if material_code != material_code2 or thickness != thickness2:
                continue

            print(f"{row_num}. material {material_code} thickness {thickness}")
            dim_A = float(row[2])
            dim_B = float(row[3])
            count = int(row[4])
            edge_A1 = int(row[5])
            edge_A2 = int(row[6])
            edge_B1 = int(row[7])
            edge_B2 = int(row[8])
            panel_name = row[9]
            panel_desc = row[10]
            cnc_face_holes = row[11]
            cnc_side_holes = row[12]

            sheet.cell(row=row_num, column=2).value = dim_A
            sheet.cell(row=row_num, column=3).value = dim_B
            sheet.cell(row=row_num, column=4).value = count
            sheet.cell(row=row_num, column=5).value = None
            sheet.cell(row=row_num, column=6).value = None
            sheet.cell(row=row_num, column=7).value = 2 if edge_A1 == 1 and edge_A2 == 1 else 1 if edge_A1 == 1 or edge_A2 == 1 else 0
            sheet.cell(row=row_num, column=8).value = 2 if edge_B1 == 1 and edge_B2 == 1 else 1 if edge_B1 == 1 or edge_B2 == 1 else 0
            sheet.cell(row=row_num, column=9).value = f"{panel_name}; {panel_desc}; {cnc_face_holes}; {cnc_side_holes}"

            row_num += 1

    # Save the new Excel file
    workbook.save(output_file)
    print(f"Successfully created order file: {output_file}")


def process_elgrad_order(workbook_file, csv_file, service, model_id):
    materials = set()
    # Read data from CSV and collect all materials and their thicknesses
    with open(csv_file, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        
        # Read header row
        header_row = next(reader)
        verify_header(header_row, CSV_HEADER)

        for row in reader:
            material_code = row[0]
            thickness = row[1]
            materials.add((material_code, thickness))

    for material_code, thickness in materials:
        material = MATERIAL[service]['iveral'][material_code]
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        output_file = os.path.join(ORDER_DIR, f"{service}_{model_id}_{date_str}_{material}_{thickness}mm.xlsx")
        process_elgrad_order_with_material(service, material_code, thickness, workbook_file, csv_file, output_file)


def process_iverpan_order(workbook_file, csv_file, service, model_id):
    date_str = datetime.datetime.now().strftime("%Y%m%d")
    output_file = os.path.join(ORDER_DIR, f"{service}_{model_id}_{date_str}.xlsx")
    print(f"Creating order file: {output_file}")
    workbook = openpyxl.load_workbook(workbook_file)
    # Select the 'Narudžba' worksheet
    sheet = workbook['Narudžba']

    # Verify header
    header_row = sheet[12]
    header_values = [cell.value for cell in header_row]
    expected_header = ['R.b', 'Šifra materijala', 'Deb.', '1.Mjera ', '2. Mjera ', 'Br.', None, None, 'R.P. kantiranje desno', ' R.P. kantiranje lijevo', 'Napomena ', 'Napomena ', 'Napomena ', 'Napomena ', 'Napomena ', None]
    verify_header(header_values, expected_header)

    sheet.cell(row=2, column=8).value = CUSTOMER_NAME
    sheet.cell(row=3, column=8).value = CUSTOMER_PHONE
    sheet.cell(row=4, column=8).value = CUSTOMER_EMAIL
    sheet.cell(row=5, column=8).value = CUSTOMER_ADDRESS

    # Read data from CSV and populate the Excel sheet
    with open(csv_file, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        
        # Read header row
        header_row = next(reader)
        verify_header(header_row, CSV_HEADER)

        # Start writing from row 14
        row_num = 14
        for row in reader:
            material_code = row[0]
            thickness = row[1]
            dim_A = float(row[2])
            dim_B = float(row[3])
            count = int(row[4])
            edge_A1 = int(row[5])
            edge_A2 = int(row[6])
            edge_B1 = int(row[7])
            edge_B2 = int(row[8])
            panel_name = row[9]
            panel_desc = row[10]
            cnc_face_holes = row[11]
            cnc_side_holes = row[12]

            sheet.cell(row=row_num, column=2).value = MATERIAL[service]['iveral'][material_code]
            sheet.cell(row=row_num, column=3).value = thickness
            sheet.cell(row=row_num, column=4).value = dim_A
            sheet.cell(row=row_num, column=5).value = dim_B
            sheet.cell(row=row_num, column=6).value = count
            sheet.cell(row=row_num, column=7).value = None if edge_A1 == 0 else MATERIAL[service]['edge_banding'][material_code]
            sheet.cell(row=row_num, column=8).value = None if edge_A2 == 0 else MATERIAL[service]['edge_banding'][material_code]
            sheet.cell(row=row_num, column=9).value = None if edge_B1 == 0 else MATERIAL[service]['edge_banding'][material_code]
            sheet.cell(row=row_num, column=10).value = None if edge_B2 == 0 else MATERIAL[service]['edge_banding'][material_code]
            
            sheet.cell(row=row_num, column=11).value = panel_name
            sheet.cell(row=row_num, column=12).value = panel_desc
            sheet.cell(row=row_num, column=13).value = cnc_face_holes
            sheet.cell(row=row_num, column=14).value = cnc_side_holes

            row_num += 1

    # Save the new Excel file
    workbook.save(output_file)
    print(f"Successfully created order file: {output_file}")


def process_furnir_order(workbook_file, csv_file, service, model_id):
    date_str = datetime.datetime.now().strftime("%Y%m%d")
    output_file = os.path.join(ORDER_DIR, f"{service}_{model_id}_{date_str}.xlsx")
    print(f"Creating order file: {output_file}")
    workbook = openpyxl.load_workbook(workbook_file)
    # Select the 'Narudžba' worksheet
    sheet = workbook['ELEMENTI']

    # Verify header
    header_row = sheet[9]
    header_values = [cell.value for cell in header_row]
    expected_header = ['Red.br.:', 'šifra1', 'šifra2', 'šifra3', 'Naziv elementa', 'CNC_PROG_1', 'CNC_PROG_2', 'NAZIV DASKE', 'Duljina  mm', 'Širina  mm', 'Debljina   mm', 'Komada', 'Materijal', 'DULJINA STRAGA', 'ŠIRINA GORE', 'DULJINA PREDNJA', 'ŠIRINA DOLJE', 'NAPOMENE', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]
    verify_header(header_values, expected_header)

    sheet.cell(row=1, column=14).value = CUSTOMER_NAME
    sheet.cell(row=3, column=14).value = CUSTOMER_EMAIL
    sheet.cell(row=4, column=14).value = CUSTOMER_PHONE

    # Read data from CSV and populate the Excel sheet
    with open(csv_file, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        
        # Read header row
        header_row = next(reader)
        verify_header(header_row, CSV_HEADER)

        # Start writing from row 14
        row_num = 10
        for row in reader:
            material_code = row[0]
            thickness = row[1]
            dim_A = float(row[2])
            dim_B = float(row[3])
            count = int(row[4])
            edge_A1 = int(row[5])
            edge_A2 = int(row[6])
            edge_B1 = int(row[7])
            edge_B2 = int(row[8])
            panel_name = row[9]
            panel_desc = row[10]
            cnc_face_holes = row[11]
            cnc_side_holes = row[12]

            sheet.cell(row=row_num, column=2).value = MATERIAL[service]['iveral'][material_code]
            sheet.cell(row=row_num, column=5).value = panel_name
            sheet.cell(row=row_num, column=6).value = cnc_face_holes
            sheet.cell(row=row_num, column=7).value = cnc_side_holes

            sheet.cell(row=row_num, column=9).value = dim_A
            sheet.cell(row=row_num, column=10).value = dim_B
            sheet.cell(row=row_num, column=11).value = thickness
            sheet.cell(row=row_num, column=12).value = count

            sheet.cell(row=row_num, column=13).value = MATERIAL[service]['iveral'][material_code]
            sheet.cell(row=row_num, column=14).value = None if edge_A1 == 0 else MATERIAL[service]['edge_banding'][material_code]
            sheet.cell(row=row_num, column=15).value = None if edge_A2 == 0 else MATERIAL[service]['edge_banding'][material_code]
            sheet.cell(row=row_num, column=16).value = None if edge_B1 == 0 else MATERIAL[service]['edge_banding'][material_code]
            sheet.cell(row=row_num, column=17).value = None if edge_B2 == 0 else MATERIAL[service]['edge_banding'][material_code]
            
            sheet.cell(row=row_num, column=18).value = panel_desc

            row_num += 1

    # Save the new Excel file
    workbook.save(output_file)
    print(f"Successfully created order file: {output_file}")


def process_sizekupres_order(workbook_file, csv_file, service, model_id):
    date_str = datetime.datetime.now().strftime("%Y%m%d")
    output_file = os.path.join(ORDER_DIR, f"{service}_{model_id}_{date_str}.xlsx")
    print(f"Creating order file: {output_file}")
    workbook = openpyxl.load_workbook(workbook_file)
    # Select the 'Narudžba' worksheet
    sheet = workbook['Iverali']

    # Verify header
    header_row = sheet[2]
    header_values = [cell.value for cell in header_row]
    expected_header = ['Iverali, iverice, lesomali, medijapan, metakrilat', 'Dužina', 'Širina', 'Količina', 'Kant', 'Napomena', None]
    verify_header(header_values, expected_header)

    # Read data from CSV and populate the Excel sheet
    with open(csv_file, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        
        # Read header row
        header_row = next(reader)
        verify_header(header_row, CSV_HEADER)

        # Start writing from row 3
        row_num = 3
        for row in reader:
            material_code = row[0]
            thickness = int(row[1])
            dim_A = float(row[2])
            dim_B = float(row[3])
            count = int(row[4])
            edge_A1 = int(row[5])
            edge_A2 = int(row[6])
            edge_B1 = int(row[7])
            edge_B2 = int(row[8])
            panel_name = row[9]
            panel_desc = row[10]
            cnc_face_holes = row[11]
            cnc_side_holes = row[12]

            sheet.cell(row=row_num, column=1).value = MATERIAL[service]['iveral'][material_code]
            sheet.cell(row=row_num, column=2).value = dim_A
            sheet.cell(row=row_num, column=3).value = dim_B
            sheet.cell(row=row_num, column=4).value = count

            # 'MEL-19': ABS 34140 RV 23/0,8 ; ABS 34140 RV 23/2
            # 'MEL-12': ABS 1615 SF 23/0,45 ; ABS 1615 SF 23/2 
            # Calculate banding (hack based on the thickness value; 19 -> 1 and 2, 12 -> 3 and 4)
            banding = lambda x: (3 if thickness <= 12 else 2) if x != 0 else 0
            banding_code = f"{banding(edge_A1)}{banding(edge_A2)}{banding(edge_B1)}{banding(edge_B2)}"
            sheet.cell(row=row_num, column=5).value = banding_code            
            sheet.cell(row=row_num, column=6).value = f"{panel_name}; {panel_desc}; {cnc_face_holes}; {cnc_side_holes}"

            row_num += 1

    # Save the new Excel file
    workbook.save(output_file)
    print(f"Successfully created order file: {output_file}")


def main():
    """Main function to orchestrate order creation."""
    # Construct file paths
    csv_file = os.path.join('export', args.model_id, 'cut_list.csv')

    # Create the output directory if it does not exist
    if not os.path.exists(ORDER_DIR):
        os.makedirs(ORDER_DIR)

    if args.service == 'iverpan':
        process_iverpan_order(args.template, csv_file, args.service, args.model_id)
    elif args.service == 'elgrad':
        process_elgrad_order(args.template, csv_file, args.service, args.model_id)
    elif args.service == 'furnir':
        process_furnir_order(args.template, csv_file, args.service, args.model_id)
    elif args.service == 'sizekupres':
        process_sizekupres_order(args.template, csv_file, args.service, args.model_id)
    else:
        print(f"Unknown service: {args.service}")
        exit(1)

if __name__ == "__main__":
    main()
